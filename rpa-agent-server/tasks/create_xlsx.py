"""
create_xlsx.py — 使用 openpyxl 生成标准报价单 .xlsx

输入 taskParams:
  deliverableId : str        — 对应 LTC-OS deliverables 表的 ID
  title         : str        — 报价单标题（e.g. "VR安全培训系统报价单"）
  customerName  : str        — 客户名称
  companyName   : str (可选) — 报价方公司名（默认"云艺化"）
  currency      : str (可选) — 货币符号（默认"¥"）
  validDays     : int (可选) — 报价有效期（天）
  deliveryWeeks : int (可选) — 交付周期（周）
  paymentTerms  : str (可选) — 付款方式
  rows          : list[dict] — 明细行
    每行: { product, qty, unit, unitPrice, total, note }
  subtotal      : float      — 小计
  discountRate  : float      — 折扣率（0~1，e.g. 0.9 = 九折）
  finalPrice    : float      — 最终成交价

输出:
  { fileUrl: "/files/xxx.xlsx" }
"""

import os
import uuid
from typing import Any
from datetime import date

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


STORAGE_DIR = os.path.join(os.path.dirname(__file__), "..", "storage")

# 颜色
COLOR_HEADER_BG = "1A3A6B"   # 深蓝（表头）
COLOR_HEADER_FG = "FFFFFF"   # 白字
COLOR_ACCENT_BG = "E8F4FD"   # 浅蓝（隔行）
COLOR_TOTAL_BG  = "008BD8"   # 天蓝（合计行）
COLOR_TOTAL_FG  = "FFFFFF"

THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE,
)


def _h(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _fmt_money(value: float, symbol: str = "¥") -> str:
    return f"{symbol}{value:,.2f}"


async def create_xlsx_task(task_params: dict[str, Any]) -> dict[str, Any]:
    os.makedirs(STORAGE_DIR, exist_ok=True)

    deliverable_id = task_params.get("deliverableId", str(uuid.uuid4()))
    title         = task_params.get("title", "报价单")
    customer_name = task_params.get("customerName", "（客户名称）")
    company_name  = task_params.get("companyName", "云艺化科技")
    currency      = task_params.get("currency", "¥")
    valid_days    = task_params.get("validDays", 30)
    delivery_weeks = task_params.get("deliveryWeeks", 8)
    payment_terms = task_params.get("paymentTerms", "首付30%，验收后付尾款70%")
    rows: list[dict] = task_params.get("rows", [])
    subtotal      = float(task_params.get("subtotal", 0))
    discount_rate = float(task_params.get("discountRate", 1.0))
    final_price   = float(task_params.get("finalPrice", subtotal * discount_rate))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报价明细"

    # --- 列宽 ---
    col_widths = [6, 30, 10, 10, 16, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # === 抬头区 ===
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = Font(name="微软雅黑", size=18, bold=True, color=COLOR_HEADER_FG)
    ws[f"A{row}"].fill = _h(COLOR_HEADER_BG)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 40
    row += 1

    # 元信息行
    meta = [
        ("报价方", company_name),
        ("客户方", customer_name),
        ("报价日期", date.today().strftime("%Y-%m-%d")),
        ("有效期", f"{valid_days} 天"),
    ]
    for label, value in meta:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="微软雅黑", size=10, bold=True)
        ws[f"A{row}"].fill = _h("F0F4FA")
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = Font(name="微软雅黑", size=10)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # 空行

    # === 表头 ===
    headers = ["序号", "产品/服务名称", "数量", "单位", "单价", "总价", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color=COLOR_HEADER_FG)
        cell.fill = _h(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    # === 明细行 ===
    for idx, item in enumerate(rows, 1):
        bg = "FFFFFF" if idx % 2 != 0 else COLOR_ACCENT_BG
        values = [
            idx,
            item.get("product", ""),
            item.get("qty", 1),
            item.get("unit", "套"),
            item.get("unitPrice", 0),
            item.get("total", 0),
            item.get("note", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.fill = _h(bg)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 3, 4) else "left", vertical="center")
            # 金额格式
            if col_idx in (5, 6) and isinstance(val, (int, float)):
                cell.number_format = f'"{currency}"#,##0.00'
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # 空行

    # === 小计行 ===
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "小计（税前）"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws[f"F{row}"] = subtotal
    ws[f"F{row}"].number_format = f'"{currency}"#,##0.00'
    ws[f"F{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    row += 1

    # 折扣行
    discount_pct = int((1 - discount_rate) * 100)
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"优惠折扣（{discount_pct}%）" if discount_pct > 0 else "折扣"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=10)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws[f"F{row}"] = -(subtotal * (1 - discount_rate))
    ws[f"F{row}"].number_format = f'"{currency}"#,##0.00'
    ws[f"F{row}"].font = Font(name="微软雅黑", size=10, color="CC0000")
    row += 1

    # 成交价（合计行，蓝底白字）
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "成交总价"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=13, bold=True, color=COLOR_TOTAL_FG)
    ws[f"A{row}"].fill = _h(COLOR_TOTAL_BG)
    ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"F{row}"] = final_price
    ws[f"F{row}"].number_format = f'"{currency}"#,##0.00'
    ws[f"F{row}"].font = Font(name="微软雅黑", size=13, bold=True, color=COLOR_TOTAL_FG)
    ws[f"F{row}"].fill = _h(COLOR_TOTAL_BG)
    ws.row_dimensions[row].height = 28
    row += 2

    # === 条款区 ===
    terms = [
        ("交付周期", f"合同签订后 {delivery_weeks} 周内完成交付"),
        ("付款方式", payment_terms),
        ("质保期",   "自验收之日起 12 个月"),
        ("备注",     "报价不含增值税，如需开票请另行协商税率"),
    ]
    for label, value in terms:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="微软雅黑", size=10, bold=True, color=COLOR_HEADER_BG)
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = Font(name="微软雅黑", size=10)
        row += 1

    # 保存文件
    filename = f"{deliverable_id}.xlsx"
    filepath = os.path.join(STORAGE_DIR, filename)
    wb.save(filepath)

    return {"fileUrl": f"/files/{filename}"}
